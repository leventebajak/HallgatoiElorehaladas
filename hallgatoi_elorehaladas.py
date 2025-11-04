import json
import logging
import os
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox, filedialog

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class CourseManagerApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Hallgatói előrehaladás kimutatás")
        self.root.geometry("600x600")

        self.courses = []  # List of tuples: (course_code, grading_type)
        self.logger = None
        self.log_file_path = None

        self.course_tree = None
        self.course_code_entry = None
        self.selected_file_var = None
        self.status_message_var = None

        self.setup_ui()

    def run(self):
        """Start the application main loop."""
        self.root.mainloop()

    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Course import/export buttons
        course_buttons_frame = ttk.Frame(main_frame)
        course_buttons_frame.grid(row=0, column=0, pady=(0, 10))

        import_button = ttk.Button(course_buttons_frame, text="Kurzusok importálása", command=self.import_from_json)
        import_button.grid(row=0, column=0, padx=(0, 5))

        export_button = ttk.Button(course_buttons_frame, text="Kurzusok exportálása", command=self.export_to_json)
        export_button.grid(row=0, column=1, padx=(5, 0))

        # Course input section
        input_frame = ttk.LabelFrame(main_frame, text="Új tárgy hozzáadása", padding="10")
        input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # Course code
        ttk.Label(input_frame, text="Tárgykód:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.course_code_entry = ttk.Entry(input_frame, width=30)
        self.course_code_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))

        # Grading type
        ttk.Label(input_frame, text="Bejegyzés típusa:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.grading_type_var = tk.StringVar(value="Évközi jegy")
        grading_combo = ttk.Combobox(
            input_frame,
            textvariable=self.grading_type_var,
            values=["Évközi jegy", "Aláírás és Vizsgajegy", "Aláírás", "Szigorlat"],
            state="readonly",
            width=28
        )
        grading_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))

        # Add button
        add_button = ttk.Button(input_frame, text="Hozzáadás", command=self.add_course)
        add_button.grid(row=2, column=0, columnspan=2, pady=(10, 0))

        input_frame.columnconfigure(1, weight=1)

        # Course list section
        list_frame = ttk.LabelFrame(main_frame, text="Hozzáadott tárgyak", padding="10")
        list_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        # Treeview for courses
        columns = ("Tárgykód", "Bejegyzés típusa")
        self.course_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        self.course_tree.heading("Tárgykód", text="Tárgykód")
        self.course_tree.heading("Bejegyzés típusa", text="Bejegyzés típusa")
        self.course_tree.column("Tárgykód", width=150)
        self.course_tree.column("Bejegyzés típusa", width=200)
        self.course_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.course_tree.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.course_tree.configure(yscrollcommand=scrollbar.set)

        # Remove button
        remove_button = ttk.Button(list_frame, text="Kijelölt törlése", command=self.remove_course)
        remove_button.grid(row=1, column=0, columnspan=2, pady=(10, 0))

        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        # File picker section
        file_picker_frame = ttk.LabelFrame(main_frame, text="Hallgatói adatok betöltése", padding="10")
        file_picker_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        self.selected_file_var = tk.StringVar(value="Nincs fájl kiválasztva")
        file_label = ttk.Label(file_picker_frame, textvariable=self.selected_file_var, relief=tk.SUNKEN, anchor=tk.W)
        file_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        browse_button = ttk.Button(file_picker_frame, text="Tallózás...", command=self.browse_excel_file)
        browse_button.grid(row=0, column=1)

        file_picker_frame.columnconfigure(0, weight=1)

        # Export button
        excel_export_button = ttk.Button(main_frame, text="Excel fájl létrehozása", command=self.export_to_excel)
        excel_export_button.grid(row=4, column=0, pady=(0, 5))

        # Status message label
        self.status_message_var = tk.StringVar(value="")
        status_label = ttk.Label(main_frame, textvariable=self.status_message_var, foreground="green", anchor=tk.CENTER)
        status_label.grid(row=5, column=0, pady=(0, 10))

        # Configure main frame
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Bind Enter key to add course
        self.course_code_entry.bind('<Return>', lambda e: self.add_course())

    def setup_logger(self, output_file_path):
        """Setup logger for the current export operation."""
        # Create log file path based on output file
        base_name = os.path.splitext(output_file_path)[0]
        self.log_file_path = f"{base_name}_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

        # Setup logger
        self.logger = logging.getLogger('CourseManager')
        self.logger.setLevel(logging.DEBUG)

        # Remove existing handlers
        self.logger.handlers.clear()

        # File handler
        file_handler = logging.FileHandler(self.log_file_path, mode='w', encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)

        # Formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)

        self.logger.addHandler(file_handler)

        self.logger.info("=== Excel fájl generálás megkezdve ===")
        self.logger.info(f"Kimeneti fájl: {output_file_path}")

    def close_logger(self):
        """Close logger and remove log file if no errors occurred."""
        if self.logger:
            # Check if any errors were logged
            has_errors = False
            if os.path.exists(self.log_file_path):
                with open(self.log_file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    has_errors = 'ERROR' in content or 'WARNING' in content

            # Close all handlers
            for handler in self.logger.handlers[:]:
                handler.close()
                self.logger.removeHandler(handler)

            # Remove log file if no errors
            if not has_errors and os.path.exists(self.log_file_path):
                try:
                    os.remove(self.log_file_path)
                    self.log_file_path = None
                except:
                    pass

            self.logger = None

    def browse_excel_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel fájlok", "*.xlsx *.xls"), ("Minden fájl", "*.*")],
            title="Hallgatói adatok kiválasztása"
        )

        if file_path:
            # Validate the Excel file
            if self.validate_excel_file(file_path):
                self.selected_file_var.set(file_path)
            else:
                self.selected_file_var.set("Nincs fájl kiválasztva")
        else:
            self.selected_file_var.set("Nincs fájl kiválasztva")

    def validate_excel_file(self, file_path):
        """Validate that the Excel file contains the required columns."""
        required_columns = [
            "Modulkód", "Felvétel féléve", "Neptun kód", "Nyomtatási név",
            "Tárgykód", "Tárgynév", "Bejegyzés értéke", "Bejegyzés típusa",
            "Bejegyzés dátuma", "Érvényes", "Elismert"
        ]

        try:
            # Read the Excel file with pandas
            df = pd.read_excel(file_path)

            # Check if all required columns are present
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                error_msg = f"A fájl nem tartalmazza a következő kötelező oszlopokat:\n" + "\n".join(missing_columns)
                messagebox.showerror("Érvénytelen fájl", error_msg)
                if self.logger:
                    self.logger.error(f"Hiányzó oszlopok: {', '.join(missing_columns)}")
                return False

            if self.logger:
                self.logger.info(f"Fájl validálva: {file_path}")
                self.logger.info(f"Sorok száma: {len(df)}")

            return True

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a fájl ellenőrzése közben:\n{str(e)}")
            if self.logger:
                self.logger.error(f"Fájl validálási hiba: {str(e)}")
            return False

    def add_course(self):
        course_code = self.course_code_entry.get().strip()
        grading_type = self.grading_type_var.get()

        if not course_code:
            messagebox.showwarning("Figyelmeztetés", "Kérem adjon meg tárgykódot!")
            return

        # Add to list
        self.courses.append((course_code, grading_type))

        # Add to treeview
        self.course_tree.insert("", tk.END, values=(course_code, grading_type))

        # Clear entry
        self.course_code_entry.delete(0, tk.END)
        self.course_code_entry.focus()

    def remove_course(self):
        selected_item = self.course_tree.selection()
        if not selected_item:
            messagebox.showwarning("Figyelmeztetés", "Kérem válasszon ki egy tárgyat a törléshez!")
            return

        # Get index
        index = self.course_tree.index(selected_item[0])

        # Remove from list
        del self.courses[index]

        # Remove from treeview
        self.course_tree.delete(selected_item)

    def import_from_json(self):
        file_path = filedialog.askopenfilename(
            defaultextension=".json",
            filetypes=[("JSON fájlok", "*.json"), ("Minden fájl", "*.*")],
            title="JSON importálása"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Validate data structure
            if not isinstance(data, list):
                messagebox.showerror("Hiba", "Érvénytelen JSON formátum. A fájlnak tárgyak listáját kell tartalmaznia.")
                return

            # Clear existing courses
            if self.courses and messagebox.askyesno("Megerősítés", "A meglévő tárgyak felülíródnak. Folytatja?"):
                self.courses.clear()
                for item in self.course_tree.get_children():
                    self.course_tree.delete(item)
            elif self.courses:
                return

            # Load courses
            for course in data:
                if isinstance(course, dict) and "course_code" in course and "grading_type" in course:
                    course_code = course["course_code"]
                    grading_type = course["grading_type"]

                    # Validate grading type
                    if grading_type not in ["Évközi jegy", "Aláírás és Vizsgajegy", "Aláírás", "Szigorlat"]:
                        continue

                    self.courses.append((course_code, grading_type))
                    self.course_tree.insert("", tk.END, values=(course_code, grading_type))

        except json.JSONDecodeError:
            messagebox.showerror("Hiba", "A fájl nem érvényes JSON formátumú.")
        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt az importálás közben:\n{str(e)}")

    def export_to_json(self):
        if not self.courses:
            messagebox.showwarning("Figyelmeztetés", "Nincs exportálandó tárgy!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON fájlok", "*.json"), ("Minden fájl", "*.*")],
            title="JSON exportálása"
        )

        if not file_path:
            return

        try:
            data = [
                {
                    "course_code": course_code,
                    "grading_type": grading_type
                }
                for course_code, grading_type in self.courses
            ]

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt az exportálás közben:\n{str(e)}")

    def export_to_excel(self):
        if not self.courses:
            messagebox.showwarning("Figyelmeztetés", "Nincs hozzáadott tárgy!")
            return

        # Check if student data file is selected
        student_file = self.selected_file_var.get()
        if student_file == "Nincs fájl kiválasztva":
            messagebox.showwarning("Figyelmeztetés", "Kérem válasszon ki egy hallgatói adatokat tartalmazó fájlt!")
            return

        # Clear previous status message
        self.status_message_var.set("")

        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel fájlok", "*.xlsx"), ("Minden fájl", "*.*")],
            title="Mentés másként"
        )

        if not file_path:
            return

        # Setup logger
        self.setup_logger(file_path)

        try:
            self.logger.info(f"Kurzusok száma: {len(self.courses)}")
            self.logger.info(f"Hallgatói adatok forrása: {student_file}")

            self.create_excel_file(file_path, student_file)

            self.logger.info("=== Excel fájl generálás sikeresen befejezve ===")

        except Exception as e:
            if self.logger:
                self.logger.error(f"Kritikus hiba az Excel fájl létrehozása során: {str(e)}", exc_info=True)
            messagebox.showerror("Hiba", f"Hiba történt a fájl létrehozása közben:\n{str(e)}")
            self.status_message_var.set("")
        finally:
            # Close logger and check for log file
            self.close_logger()

            # Update status message after logger is closed
            if self.log_file_path and os.path.exists(self.log_file_path):
                # There were errors/warnings - show in orange/red
                self.status_message_var.set(
                    f"⚠ Excel fájl létrehozva, de vannak figyelmeztetések!\nLog fájl: {self.log_file_path}")
                # Change label color to orange for warnings
                for widget in self.root.winfo_children():
                    self._update_status_label_color(widget, "orange")
            else:
                # Success - show in green
                self.status_message_var.set(f"✓ Excel fájl sikeresen létrehozva: {file_path}")
                # Change label color to green
                for widget in self.root.winfo_children():
                    self._update_status_label_color(widget, "green")

    def _update_status_label_color(self, widget, color):
        """Recursively find and update the status label color."""
        if isinstance(widget, ttk.Label) and widget.cget('textvariable') == str(self.status_message_var):
            widget.configure(foreground=color)
        for child in widget.winfo_children():
            self._update_status_label_color(child, color)

    def get_unique_students(self, student_file):
        """Extract unique student data from the student Excel file."""
        try:
            df = pd.read_excel(student_file)

            # Select the 4 columns and drop duplicates
            students_df = df[["Modulkód", "Felvétel féléve", "Neptun kód", "Nyomtatási név"]].drop_duplicates()

            if self.logger:
                self.logger.info(f"Egyedi hallgatók száma: {len(students_df)}")

            # Convert to list of tuples
            return [tuple(row) for row in students_df.values]
        except Exception as e:
            if self.logger:
                self.logger.error(f"Hiba az egyedi hallgatók kinyerése során: {str(e)}")
            raise

    def load_student_data(self, student_file):
        """Load all student data from the Excel file once using pandas."""
        try:
            df = pd.read_excel(student_file)

            # Select only the columns we need and keep all rows (including invalid ones for signature checking)
            df = df[["Neptun kód", "Tárgykód", "Bejegyzés értéke", "Bejegyzés dátuma", "Érvényes", "Bejegyzés típusa", "Elismert"]]

            if self.logger:
                self.logger.info(f"Hallgatói adatok betöltve: {len(df)} bejegyzés")

            return df
        except Exception as e:
            if self.logger:
                self.logger.error(f"Hiba a hallgatói adatok betöltése során: {str(e)}")
            raise

    def get_student_grade(self, student_data, neptun_kod, course_code):
        """Get the latest valid grade for a student in a specific course from cached data.
        Returns tuple: (grade, is_recognized)"""
        try:
            # Filter for the specific student, course, and valid entries
            filtered = student_data[
                (student_data["Neptun kód"] == neptun_kod) &
                (student_data["Tárgykód"] == course_code) &
                (student_data["Érvényes"].isin(["Igaz", True, "TRUE", "true"]))
                ]

            if filtered.empty:
                if self.logger:
                    self.logger.debug(f"Nincs érvényes bejegyzés - Neptun: {neptun_kod}, Tárgykód: {course_code}")
                return None, False

            # Sort by date and get the most recent grade
            filtered = filtered.sort_values("Bejegyzés dátuma", ascending=False)
            grade = filtered.iloc[0]["Bejegyzés értéke"]
            is_recognized = filtered.iloc[0]["Elismert"] in ["Igaz", True, "TRUE", "true"]

            if self.logger:
                self.logger.debug(f"Jegy talált - Neptun: {neptun_kod}, Tárgykód: {course_code}, Jegy: {grade}, Elismert: {is_recognized}")

            return grade, is_recognized
        except Exception as e:
            if self.logger:
                self.logger.warning(
                    f"Hiba a jegy lekérése során - Neptun: {neptun_kod}, Tárgykód: {course_code}: {str(e)}")
            return None, False

    def get_student_signature_and_exam(self, student_data, neptun_kod, course_code):
        """Get signature and exam grade for a student in a specific course.
        Returns tuple: (signature, exam_grade, signature_is_recognized, exam_is_recognized)"""
        try:
            # Filter for the specific student and course
            filtered = student_data[
                (student_data["Neptun kód"] == neptun_kod) &
                (student_data["Tárgykód"] == course_code)
                ]

            if filtered.empty:
                if self.logger:
                    self.logger.debug(f"Nincs bejegyzés - Neptun: {neptun_kod}, Tárgykód: {course_code}")
                return None, None, False, False

            # Filter for Aláírás entries
            alairas = filtered[filtered["Bejegyzés típusa"] == "Aláírás"]

            if alairas.empty:
                if self.logger:
                    self.logger.warning(f"Nincs aláírás bejegyzés - Neptun: {neptun_kod}, Tárgykód: {course_code}")
                return None, None, False, False

            # Sort by date and get the most recent signature
            alairas = alairas.sort_values("Bejegyzés dátuma", ascending=False)
            latest_signature = alairas.iloc[0]["Bejegyzés értéke"]
            signature_is_recognized = alairas.iloc[0]["Elismert"] in ["Igaz", True, "TRUE", "true"]

            # If signature is "Megtagadva", return it with no exam grade
            if latest_signature == "Megtagadva":
                if self.logger:
                    self.logger.debug(f"Aláírás megtagadva - Neptun: {neptun_kod}, Tárgykód: {course_code}")
                return latest_signature, None, signature_is_recognized, False

            # If signature is "Aláírva", look for exam grade
            if latest_signature == "Aláírva":
                # Filter for Vizsgajegy entries that are valid
                vizsgajegy = filtered[
                    (filtered["Bejegyzés típusa"] == "Vizsgajegy") &
                    (filtered["Érvényes"].isin(["Igaz", True, "TRUE", "true"]))
                    ]

                if not vizsgajegy.empty:
                    # Sort by date and get the most recent exam grade
                    vizsgajegy = vizsgajegy.sort_values("Bejegyzés dátuma", ascending=False)
                    exam_grade = vizsgajegy.iloc[0]["Bejegyzés értéke"]
                    exam_is_recognized = vizsgajegy.iloc[0]["Elismert"] in ["Igaz", True, "TRUE", "true"]
                    if self.logger:
                        self.logger.debug(
                            f"Aláírás és Vizsgajegy - Neptun: {neptun_kod}, Tárgykód: {course_code}, Jegy: {exam_grade}, Elismert: {exam_is_recognized}")
                    return latest_signature, exam_grade, signature_is_recognized, exam_is_recognized
                else:
                    if self.logger:
                        self.logger.debug(
                            f"Aláírva, de nincs vizsgajegy - Neptun: {neptun_kod}, Tárgykód: {course_code}")
                    return latest_signature, None, signature_is_recognized, False

            # For any other signature value
            if self.logger:
                self.logger.debug(
                    f"Egyéb aláírás érték - Neptun: {neptun_kod}, Tárgykód: {course_code}, Érték: {latest_signature}")
            return latest_signature, None, signature_is_recognized, False
        except Exception as e:
            if self.logger:
                self.logger.warning(
                    f"Hiba az aláírás/vizsgajegy lekérése során - Neptun: {neptun_kod}, Tárgykód: {course_code}: {str(e)}")
            return None, None, False, False

    def create_excel_file(self, file_path, student_file):
        if self.logger:
            self.logger.info("Excel fájl létrehozás megkezdve")

        # Load all student data once
        student_data = self.load_student_data(student_file)

        # Get unique student data
        students = self.get_unique_students(student_file)

        wb = Workbook()
        ws = wb.active
        ws.title = "Tantárgyak"

        if self.logger:
            self.logger.info("Munkafüzet létrehozva")

        # Define colors
        colors = ["D9E1F2", "B4C6E7"]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define bold font for headers
        bold_font = Font(bold=True)

        # Define number of data rows based on student count
        num_data_rows = len(students) + 1  # +1 for header row

        # Base columns
        base_headers = [
            "Modulkód",
            "Felvétel féléve",
            "Neptun kód",
            "Nyomtatási név",
            "Felvételi összes pontszám",
            "Státusz"
        ]

        # Add base headers and apply borders to entire columns
        for col_idx, header in enumerate(base_headers, start=1):
            # Header cell
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = bold_font
            cell.border = thin_border

            # Apply borders to all rows in this column
            for row_idx in range(2, num_data_rows + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

        # Fill in student data (first 4 columns)
        if self.logger:
            self.logger.info("Hallgatói alapadatok kitöltése...")

        for row_idx, student in enumerate(students, start=2):
            ws.cell(row=row_idx, column=1, value=student[0])  # Modulkód
            ws.cell(row=row_idx, column=2, value=student[1])  # Felvétel féléve
            ws.cell(row=row_idx, column=3, value=student[2])  # Neptun kód
            ws.cell(row=row_idx, column=4, value=student[3])  # Nyomtatási név

        current_col = len(base_headers) + 1
        color_idx = 0

        # Add course columns
        if self.logger:
            self.logger.info("Kurzusok feldolgozása...")

        for course_code, grading_type in self.courses:
            if self.logger:
                self.logger.info(f"Kurzus feldolgozása: {course_code} ({grading_type})")

            color = colors[color_idx % len(colors)]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            if grading_type == "Aláírás és Vizsgajegy":
                # Merge two columns for header
                ws.merge_cells(
                    start_row=1,
                    start_column=current_col,
                    end_row=1,
                    end_column=current_col + 1
                )

                # Set header value and formatting
                cell = ws.cell(row=1, column=current_col, value=course_code)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill
                cell.font = bold_font
                cell.border = thin_border

                # Apply fill and border to the second merged cell in header
                second_cell = ws.cell(row=1, column=current_col + 1)
                second_cell.fill = fill
                second_cell.border = thin_border

                # Apply borders and background to all rows in both columns, and fill in signature/exam data
                for row_idx in range(2, num_data_rows + 1):
                    cell1 = ws.cell(row=row_idx, column=current_col)
                    cell1.border = thin_border
                    cell1.fill = fill

                    cell2 = ws.cell(row=row_idx, column=current_col + 1)
                    cell2.border = thin_border
                    cell2.fill = fill

                    # Get the student's Neptun code from column 3
                    neptun_kod = ws.cell(row=row_idx, column=3).value
                    if neptun_kod:
                        try:
                            # Get the signature and exam grade for this student and course
                            signature, exam, sig_recognized, exam_recognized = self.get_student_signature_and_exam(student_data, neptun_kod, course_code)
                            if signature:
                                cell1.value = signature
                            else:
                                # Student didn't take the course - color both cells yellow
                                cell1.fill = yellow_fill
                                cell2.fill = yellow_fill
                            if exam:
                                cell2.value = exam
                            # If recognized, color green, otherwise keep the alternating color
                            if sig_recognized and exam_recognized:
                                cell1.fill = green_fill
                                cell2.fill = green_fill
                        except Exception as e:
                            if self.logger:
                                self.logger.error(
                                    f"Hiba a {neptun_kod} hallgató {course_code} kurzusának feldolgozása során: {str(e)}")

                current_col += 2
            else:  # "Évközi jegy", "Szigorlat", or "Aláírás"
                # Single column header
                cell = ws.cell(row=1, column=current_col, value=course_code)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill
                cell.font = bold_font
                cell.border = thin_border

                # Apply borders and background to all rows in this column, and fill in grades
                for row_idx in range(2, num_data_rows + 1):
                    cell = ws.cell(row=row_idx, column=current_col)
                    cell.border = thin_border
                    cell.fill = fill

                    # Get the student's Neptun code from column 3
                    neptun_kod = ws.cell(row=row_idx, column=3).value
                    if neptun_kod:
                        try:
                            # Get the grade/signature for this student and course from cached data
                            grade, is_recognized = self.get_student_grade(student_data, neptun_kod, course_code)
                            if grade:
                                cell.value = grade
                                # If recognized, color green, otherwise keep the alternating color
                                if is_recognized:
                                    cell.fill = green_fill
                            else:
                                # Student didn't take the course - color cell yellow
                                cell.fill = yellow_fill
                        except Exception as e:
                            if self.logger:
                                self.logger.error(
                                    f"Hiba a {neptun_kod} hallgató {course_code} kurzusának feldolgozása során: {str(e)}")

                current_col += 1

            color_idx += 1

        if self.logger:
            self.logger.info("Oszlopszélességek beállítása...")

        # Auto-adjust column widths
        # For base columns, adjust based on header content
        for col_idx in range(1, len(base_headers) + 1):
            column_letter = get_column_letter(col_idx)
            header_length = len(base_headers[col_idx - 1])
            # Add some padding (minimum 2 extra characters)
            ws.column_dimensions[column_letter].width = max(header_length + 2, 10)

        # For course columns, set fixed width
        for col_idx in range(len(base_headers) + 1, current_col):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20

        # Save workbook
        if self.logger:
            self.logger.info(f"Munkafüzet mentése: {file_path}")

        wb.save(file_path)

        if self.logger:
            self.logger.info("Munkafüzet sikeresen mentve")


def main():
    app = CourseManagerApp()
    app.run()


if __name__ == "__main__":
    main()
